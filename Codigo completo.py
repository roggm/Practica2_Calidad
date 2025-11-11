import openpyxl

import Funcionesproy

libro = openpyxl.Workbook()
nombrehoja = input("Nombre de la hoja: ")
hoja1 = libro.create_sheet(nombrehoja)
#predeterminedSheets = libro.get_sheet_names()
std = libro["Sheet"]
hoja1.cell(1,1).value= "Valores"
hoja1.cell(2,1).value= "Alumnos"
print("¿Cuántos criterios de calificación quieres?")
nCriterios = int(input ("Número de criterios a agregar: "))
nAlumnos = int(input ("¿Cuántos alumnos quieres registrar? "))
Funcionesproy.crearCriteriosYAlumnos(nCriterios,hoja1,nAlumnos)
libro.remove(std)
libro.save("Calculadora.xlsx")

